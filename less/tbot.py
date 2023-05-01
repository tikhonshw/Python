import requests
import schedule
import os
import subprocess
import telebot
from telebot import types
import markdown
import pyowm
import random
import time
from multiprocessing.context import Process
import math
from time import sleep
from threading import Thread
import textwrap

# from datetime import datetime
from datetime import datetime, timedelta
# import timedelta
import mysql.connector

# для обработки xlsx
from openpyxl import load_workbook

## для ямальской погоды 
from bs4 import BeautifulSoup as bs4

bot = telebot.TeleBot("5125341770:AAF11nLzMCoeFV-gf96iL19hDyhOfidqo7g")

#
# from pyowm import OWM
# owm = OWM('a58cd71e6d3057b4ce76c3da27076585')
# mgr = owm.weather_manager()
# observation = mgr.weather_at_place('Москва')
# w = observation.weather
# print(w, w.wind())


# сохранение запросов в базу ###################################
def saveTbotLog(msg):
    # сохраняем запросы пользователей 
    mybdTbot = mysql.connector.connect(
        host="10.152.36.10",
        port="3306",
        user="root",
        password="3333",
        database="arm_vdo"
    )
    todayTmp = datetime.today()
    todayTmp = todayTmp.strftime("%Y-%m-%d %H:%M:%S")
    myCurTbot = mybdTbot.cursor(buffered=True) 
    sqlTbot = 'INSERT INTO tbot_log (user_id, first_name, last_name, username, text, dateAdd) VALUES (%s, %s, %s, %s, %s, %s)'
    myCurTbot.execute(sqlTbot, (msg.from_user.id, msg.from_user.first_name, msg.from_user.last_name, msg.from_user.username, msg.text, todayTmp))
    # print(message.from_user.id, message.from_user.first_name, message.from_user.last_name, message.from_user.username, message.text)
    mybdTbot.commit()
################################################################

# обработка файла для получения инофрмации кто на смене ########
def getInfoSmena(dateInfo, dolgnost, smena):
    ddmmYYYY = dateInfo.split(".")
    dd = int(ddmmYYYY[0]) * 1
    mm = int(ddmmYYYY[1]) * 1
    massMonth = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь",
                    "Октябрь", "Ноябрь", "Декабрь"]
    if dolgnost == 'tech':
        wb = load_workbook('tech.xlsx')  # подключаем файл график техников
    if dolgnost == 'ing':
        wb = load_workbook('ing.xlsx')  # подключаем файл график техников
    sheet = wb.get_sheet_by_name(massMonth[mm])
    for i in range(15, 32):
        if sheet.cell(row=i, column=dd + 1).value is not None:
            workDay = sheet.cell(row=i, column=dd + 1).value
            if str(smena) == str(workDay):
                fio = sheet.cell(row = i - 1, column = 1).value
                fio = fio.replace('.','. ')
                return fio
            # else:
            #     return 'None'
                # return sheet.cell(row=i - 1, column=1).value
################################################################

@bot.message_handler(commands=['help'])
def website(message):
    print(message)
    saveTbotLog(message)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=4)
    smena = types.KeyboardButton('Смены')
    pogoda = types.KeyboardButton('Погода')
    plan = types.KeyboardButton('План на день')
    
    pokazElektro = types.KeyboardButton('Показания ЭЭ')
    narabFazan = types.KeyboardButton('ФАЗАН 19Р')
    if message.from_user.id == 1308468425:
        vpn = types.KeyboardButton('ВПН')
        liana = types.KeyboardButton('Лиана')
    else:
        vpn = ''
        liana = ''
    help = types.KeyboardButton('Помощь')
    msgHelp = '''<b>/help@kdp_yanao_bot</b> - Информация по командам
<b>/smena@kdp_yanao_bot</b> - Информация о смене
<b>/pogoda@kdp_yanao_bot</b> - Информация о погоде
<b>/plan@kdp_yanao_bot</b> - план ТО на сегодня
'''
    bot.delete_message(message.chat.id, message.message_id)
    markup.add(smena, pogoda, plan, liana, pokazElektro, narabFazan, vpn, help)
    bot.send_message(message.from_user.id, msgHelp, reply_markup=markup, parse_mode='html') # message.chat.id



@bot.message_handler(commands=['reg'])
def reg(message):
    # print(message)
    saveTbotLog(message)

    mess = f'Привет, <b>{message.from_user.first_name} <u>{message.from_user.last_name}</u></b>'
    bot.send_message(message.from_user.id, mess, parse_mode='html') #message.chat.id

@bot.callback_query_handler(func=lambda call: True)
def callback(call):
    # print(call)
    # print('============', call.data[0:5])
    # saveTbotLog(call)
    if call.data == 'LianaMenu':
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        new_menu.add(types.InlineKeyboardButton('FreeSpace%', callback_data='lianaFreeSpace'),
                     types.InlineKeyboardButton('LastBackUp', callback_data='lianaLastBackUp'),
                     types.InlineKeyboardButton('ClearSpace', callback_data='lianaClearSpace') )
        bot.edit_message_text('Выберите действие:', call.message.chat.id, call.message.message_id, reply_markup=new_menu)
    if call.data == 'lianaFreeSpace':
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        percent = os.popen("sshpass -p otstoj111 ssh -p 10018 user@10.3.103.100 df / | grep / | awk '{ print $5}' | sed 's/%//g'").read()
        answerMess = "Корневой раздел сервера ЛИАНА использует дисковое пространство на: " + str(int(percent)) + "%"
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='LianaMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu)
    if call.data == 'lianaLastBackUp':
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        lastFile = os.popen("ls --full-time  /home/user/MySoft/###Liana/###Back-Up/ | tail -n 1 | awk '{print \"Архив от: \" $6 \" имя файла: \" $9 \" размер \" $5 / 1024 / 1024 \" Mб.\"}'").read() 
        answerMess = lastFile
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='LianaMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu)
    if call.data == 'lianaClearSpace':
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "INCOMPLETE"
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='LianaMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu)

    
    #  avr2 negarantiya kdp1 kdp2
    if call.data == 'pokazElektroMenu':
        global typeEE
        typeEE = ''
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        new_menu.add(types.InlineKeyboardButton('АВР 1', callback_data='avr1'),
                     types.InlineKeyboardButton('АВР 2', callback_data='avr2'),
                     types.InlineKeyboardButton('АВР 2', callback_data='avr2'),
                     types.InlineKeyboardButton('КДП №1', callback_data='kdp1'),
                     types.InlineKeyboardButton('КДП №2', callback_data='kdp2') )
        bot.edit_message_text('Выберите действие:', call.message.chat.id, call.message.message_id, reply_markup=new_menu)
    if call.data == 'pokazElektroMenuCancel':
        answerMess = 'Ввод показаний ЭЭ отменен'
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id) 
    if call.data == 'avr1': 
        typeEE = call.data
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "Введите показания счетчика ЭЭ АВР 1 и нажмите кнопку \"Отправить\""
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data = 'pokazElektroMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvod, deleteMsg)
    if call.data == 'avr2':
        typeEE = call.data
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "Введите показания счетчика ЭЭ АВР 2 и нажмите кнопку \"Отправить\""
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='pokazElektroMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvod, deleteMsg)
    if call.data == 'negarantiya':
        typeEE = call.data
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "Введите показания счетчика ЭЭ \"Негарантия\" и нажмите кнопку \"Отправить\""
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='pokazElektroMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvod, deleteMsg)
    if call.data == 'kdp1':
        typeEE = call.data
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "Введите показания счетчика ЭЭ \"КДП 1\" и нажмите кнопку \"Отправить\""
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='pokazElektroMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvod, deleteMsg)
    if call.data == 'kdp2':
        typeEE = call.data
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = "Введите показания счетчика ЭЭ \"КДП 2\" и нажмите кнопку \"Отправить\""
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data='pokazElektroMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvod, deleteMsg)
    if call.data == 'SendDataEE':            
        try: 
            dictNamePriborEE = {'avr1': 'АВР 1 №12254085',
                                'avr2': 'АВР 2 №12254023',
                                'negarantiya': 'Негарантия',
                                'kdp1': 'КДП 1 №28371868',
                                'kdp2': 'КДП 2 №27472779'
                                }
            answerMess = 'Приняты новые показания по ЭЭ\n'
            for key in dic: 
                answerMess += "{}: {}\n".format(dictNamePriborEE[key],dic[key])
            bot.send_message(1308468425, answerMess)  

            answerMess = 'Показания по ЭЭ успешно переданы'
            bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id) 
            dic.clear()
        except Exception: 
            answerMess = 'Неудалось отправить показания ЭЭ'
            bot.send_message(1308468425, answerMess)  

    ###### обработка фазанов
    # print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!' , call.data[1])
    if call.data == 'pokazFazanMenu':
        getZeroData()
        showMenuFazan(call.from_user.id, 1, call)
    if call.data == 'pokazFazanMenuCancel':
        answerMess = 'Ввод показаний наработки ФАЗАН 19Р отменен'
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id) 
    if call.data[0] == 'f': 
        deleteMsg = call.message.message_id
        new_menu = types.InlineKeyboardMarkup(row_width=1)
        answerMess = 'Введите наработу для ' + getInfoFazan(call.data)[0] + " №" + getInfoFazan(call.data)[1]
        new_menu.add(types.InlineKeyboardButton('<< Back', callback_data = 'pokazFazanMenu'))
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id, reply_markup=new_menu) 
        bot.register_next_step_handler_by_chat_id(call.message.chat.id, getVvodNarabotkaFazan, deleteMsg, getInfoFazan(call.data)[0], call ) 
    if call.data == 'SendDataFazan':
        try: 
            answerMess = 'Приняты новые показания по наработке ФАЗАН 19Р\n'
            for key, value in dictFazan.items():  
                answerMess += "{} №{} наработка: {}\n".format(key, value[1], value[2])                
            bot.send_message(1308468425, answerMess)  
            for key, value in dictFazan.items():  
                dictFazan[key].insert(2, '0')
            answerMess = 'Показания по наработке ФАЗАН 19Р'
            bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id)             
        except Exception: 
            answerMess = 'Неудалось отправить информацию по наработке ФАЗАН 19Р'
            bot.send_message(1308468425, answerMess) 
    
    ###### обработка команд ВПН
    def statuVpn():
        cmd_str = ''; 
        text = subprocess.run('ifconfig -a | grep 10.8.0 | awk \'{ print $2}\'', shell=True, stdout=subprocess.PIPE, encoding='utf-8')
        # text.stdout.read()[0]
        if str(text.stdout) == '':
            answerMess = 'VPN не подключен'
            bot.send_message(1308468425, answerMess) 
        else:
            answerMess = 'VPN подлючен. Адрес в сети ' + str(text.stdout)
            bot.send_message(1308468425, answerMess) 
    if call.data == 'vpnStart':
        # os.system('/home/user/MySoft/###OpenVPN/connect.sh')
        cmd_str = "/home/user/MySoft/###OpenVPN/connect.sh"        
        subprocess.run(cmd_str, shell=True)
        statuVpn()
    if call.data == 'vpnStop':
        os.system('echo "29072019" | sudo -S killall openvpn')
    if call.data == 'vpnStatus':
        statuVpn()
        
        
         
    if call.data == 'vpnExit':
        answerMess = 'Вы вышли из меню VPN'
        bot.edit_message_text(answerMess, call.message.chat.id, call.message.message_id) 





def getVvodNarabotkaFazan(message, deleteMsg, saveNameFazan, call):
    # print('--------------------------------------')
    # print('--------------------------------------')
    # print('--------------------------------------')
    # print('saveNameFazan: ' , saveNameFazan , dictFazan[saveNameFazan][2] )
    # print('--------------------------------------')
    # print('--------------------------------------')
    # print('--------------------------------------')
    
    # bot.delete_message(message.chat.id, deleteMsg)
    bot.delete_message(message.chat.id, message.message_id)  
    try:
        if message.text.isdigit() == True:
            # print('All ok')
            answerUser = message.from_user.id 
            dictFazan[saveNameFazan].insert(2, message.text)
            showMenuFazan(call.from_user.id, 1, call)
        # else: 
            # bot.send_message(message.from_user.id, 'Цифрами, пожалуйста')
            # bot.register_next_step_handler(call.message.chat.id, getVvodNarabotkaFazan, deleteMsg, saveNameFazan, call ) 
    except Exception():
        bot.send_message(message.from_user.id, 'Не удалось сохранить показания ФААЗАН ' + saveNameFazan)
        print('You send me message', message.text, 'answerUser: ', answerUser)

def getVvod(message, deleteMsg): 
    bot.delete_message(message.chat.id, deleteMsg)
    bot.delete_message(message.chat.id, message.message_id) 
    try:
        if message.text.isdigit() == True:
            answerUser = message.from_user.id 
            dic.update({typeEE: message.text})
            showMenuPokazEE(answerUser) 
        else:
            bot.send_message(message.from_user.id, 'Цифрами, пожалуйста')
            bot.register_next_step_handler(message, getVvod)
    except Exception:
        bot.send_message(message.from_user.id, 'Не удалось выполнить поставленную задачу, пробуйте еще раз')
        print('You send me message', message.text, 'typeEE', typeEE, 'answerUser: ', answerUser) 

def showMenuPokazEE(answerUser):
    # print(dic)
    buttons = types.InlineKeyboardMarkup(row_width=1)
    button1 = types.InlineKeyboardButton('АВР 1 ' + dic['avr1'], callback_data='avr1')
    button2 = types.InlineKeyboardButton('АВР 2 ' + dic['avr2'], callback_data='avr2')
    button3 = types.InlineKeyboardButton('негарантия ' + dic['negarantiya'], callback_data='negarantiya')
    button4 = types.InlineKeyboardButton('КДП №1 ' + dic['kdp1'], callback_data='kdp1')
    button5 = types.InlineKeyboardButton('КДП №2 ' + dic['kdp2'], callback_data='kdp2')
    button6 = types.InlineKeyboardButton('<< Отменить', callback_data='pokazElektroMenuCancel') 
    buttons.add(button1, button2, button3, button4, button5, button6) 
    if '' not in dic.values():
        button6 = types.InlineKeyboardButton('Отправить показания ЭЭ', callback_data='SendDataEE')
        buttons.add(button6)
    bot.send_message(answerUser, text='Выберите действие:', reply_markup=buttons) 

    
def getInfoFazan(name):
    for key, value in dictFazan.items(): 
        if value[0] == name:
            return (key, value[1])
def getZeroData():
    c = 0
    for key, value in dictFazan.items(): 
        if value[2] == '0':
            c = c + 1
    return c
        
dictFazan = {'СДП резерв':   ['f_sdp_rez',     '10320083',  '0'],
             'ЛАЗ КДП':      ['f_laz_kdp',     '10320084',  '0'],
             'ДПК АКБ':      ['f_dpk_akb',     '112151006', '0'],
             'СДП АКБ':      ['f_sdp_akb',     '112151001', '0'],
             'ЦПИ Юг акб':   ['f_cpi_yg_akb',  '11215979',  '0'],
             'Ст. дисп акб': ['f_st_disp_akb', '11215997',  '0'],
             'ДПК резерв':   ['f_dpk_rez',     '10320072',  '0']
            }
def showMenuFazan(answerUser, stat, call):
    buttons = types.InlineKeyboardMarkup(row_width=1) 
    for key, value in dictFazan.items():
        nameButton = key + ' №' + value[1]
        # print(key, value[0], value[1], value[2])
        if value[2] == '0':
            nameButton = key + ' №' + value[1] 
        else: 
            nameButton = key + ' - нараб.: ' + value[2] 
        buttons.add(
            types.InlineKeyboardButton(
                nameButton, 
                callback_data=value[0])) 
    buttons.add(types.InlineKeyboardButton('<< Отменить', callback_data='pokazFazanMenuCancel'))

    # print( getZeroData() )

    if getZeroData() == 0:
        buttons.add(types.InlineKeyboardButton('Отправить показания', callback_data='SendDataFazan')) 

    if stat == 0:
        bot.send_message(answerUser, text='Добавление наработки:', reply_markup=buttons)
    if stat == 1: 
        bot.edit_message_text('Добавление наработки ФАЗАН 19Р:', 
                              call.message.chat.id, call.message.message_id, reply_markup=buttons) 

    

@bot.message_handler(content_types=['text'])
def send_echo(message):
    saveTbotLog(message)
    # print(message)
    # print('***********************')
    # print(message.chat.title)
    # print('***********************')
    # if message.chat.id == "-718406466":
    if message.chat.title == "КДП ЯНАО" and (message.text == "План на день" or message.text == "Погода" or message.text == "Смены" or message.text == "Помощь" or message.text == "/smena@kdp_yanao_bot" or message.text == "/help@kdp_yanao_bot" or message.text == "/pogoda@kdp_yanao_bot" or message.text == "/plan@kdp_yanao_bot"): #
        print('***********************')
        print('DELETE')
        print('***********************')
        bot.delete_message(message.chat.id, message.message_id)
        answerUser = message.from_user.id
    else: 
        answerUser = message.from_user.id


    # делим строку для дальнейше обработки
    inCommingSTR = message.text
    if "@" in inCommingSTR:
        commandStr = inCommingSTR.split("@")
        commandStr = commandStr[0]
        chatStr = commandStr[1]
    else: commandStr = inCommingSTR

    if commandStr == "Показания ЭЭ": 
        global dic
        dic = {'avr1': '', 'avr2': '', 'negarantiya': '','kdp1': '','kdp2': ''}
        showMenuPokazEE(answerUser)
    
    if commandStr == "ФАЗАН 19Р": 
        global dicTimeFazan
        # dicTimeFazan = {}
        showMenuFazan(answerUser, 0, 0)
    
    if commandStr == "Лиана":
        buttons = types.InlineKeyboardMarkup(row_width=1)
        button1 = types.InlineKeyboardButton('FreeSpace%', callback_data='lianaFreeSpace')
        button2 = types.InlineKeyboardButton('LastBackUp', callback_data='lianaLastBackUp')
        button3 = types.InlineKeyboardButton('ClearSpace', callback_data='lianaClearSpace')
        buttons.add(button1, button2, button3)
        bot.send_message(answerUser, text='Выберите действие:', reply_markup=buttons) 
    
    if commandStr == "ВПН":
        buttons = types.InlineKeyboardMarkup(row_width=1)
        button1 = types.InlineKeyboardButton('Включить VPN', callback_data='vpnStart')
        button2 = types.InlineKeyboardButton('Отключить VPN', callback_data='vpnStop')
        button3 = types.InlineKeyboardButton('Статус VPN', callback_data='vpnStatus')
        
        button4 = types.InlineKeyboardButton('<< Back', callback_data='vpnExit')
        buttons.add(button1, button2, button3, button4)
        bot.send_message(answerUser, text='Выберите действие: ', reply_markup=buttons) 

    if commandStr == "Погода" or commandStr == "/pogoda":
        ### погода Ямальская ##########################################
        answerMess = 'По даным Салехард МЕТЕО:\n'
        headers = {
            'accept': '*/*',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.70 Safari/537.36'
        }
        # url = "https://itproger.com/"
        # https://meteoinfo.ru/forecasts5000/russia/jamalo-nenetskij-ar/salehard#2119344415
        url = 'https://meteoinfo.ru/rss/forecasts/index.php?s=23330'
        session = requests.Session()

        try:
            req = session.get(url, headers=headers)
            if req.status_code == 200:
                soup = bs4(req.content, 'html.parser')
                # print (soup)
                divs = soup.find_all('item')
                # print (divs)
                for div in divs:
                    title = div.find('title').text
                    href = div.find('description').text
                    answerMess += "\n{} \n {}\n".format(title, href)
                    # print("{} \n {}".format(title, href))
                # bot.send_message(message.chat.id, answerMess)
            else:
                print("Ошибка")
        except Exception:
            print("Ошибка в самом URL адресе")
        ###############################################################



        owm = pyowm.OWM('a58cd71e6d3057b4ce76c3da27076585')
        ob = owm.weather_at_place('Салехард, RU')
        w = ob.get_weather()
        temp = w.get_temperature('celsius')["temp"]
        wind = w.get_wind()["speed"]
        windGust = w.get_wind()["gust"]
        pressure = w.get_pressure()["press"] * 0.750063755419211  # / 1.333
        answerMess += "\nПо даным OpenWeather\nВ Салехарде сейчас:\n\tтемпература: " + str(temp) + "°С\tветер: " + str(
            wind) + " м/с\tпорывы до: " + str(windGust) + " м/с\tдавление: " + str(round(pressure, 1)) + " mmHg"
        # bot.send_message(message.chat.id, answerMess)

        ###########################
        ### https://wttr.in/:help?lang=ru
        # url = 'https://wttr.in'  # не изменяйте значение URL
        #
        # weather_parameters = {
        #     '0': '',
        #     'format': 4,
        #     'T': '',
        #     'M': '',
        #     'm': '',
        #     'P': 'hPa',
        #     'lang': 'ru',
        #     # добавьте параметр запроса `T`, чтобы вернулся чёрно-белый текст
        # }
        #
        # response = requests.get(url + '/Салехард', params=weather_parameters)  # передайте параметры в http-запрос
        # answerMess += '\n\nИз альтернативного источника:\n'
        # answerMess += str(response.text)
        # print(response.text)
        cities = [
            'Салехард',
            'Лабытнанги',
            'Тюмень',
            'Омск',
            'Екатеринбург',
            'Рио'
        ]
        def make_url(city):
            # в URL задаём город, в котором узнаем погоду
            return f'http://wttr.in/{city}'


        def make_parameters():
            params = {
                '0': '',
                'format': 4,
                'T': '',
                'M': '',
                'm': '',
                'P': 'hPa',
                'lang': 'ru',
            }
            return params

        def what_weather(city):
            try:
                url = make_url(city)
                www = requests.get(url + '?P', params=make_parameters())
                return www.text
            except e:
                return '<сетевая ошибка>'

        def what_weather(city):
            # Напишите тело этой функции.
            # Не изменяйте остальной код!
            try:
                url = make_url(city)
                www = requests.get(url, params=make_parameters())
                return www.text
            except e:
                return '<сетевая ошибка>'

        answerMess += '\n\nИз альтернативного источника:\n'
        for city in cities:
            answerMess += what_weather(city)
        # print('Погода в городах:')


        ###########################

        bot.send_message(answerUser, answerMess)

    # elif commandStr == "/smena":
    elif commandStr == "Смены" or commandStr == "/smena":
        dateStr = False
        if " " in commandStr:
            commandStrTemp = commandStr.split(" ")
            # print(commandStr[0] + " hui " + commandStr[1])
            if commandStrTemp[0] == "/smena":
                commandStr = commandStrTemp[0]
                dateStr = str(commandStrTemp[1])
                # print(str(commandStr) + " :: " + str(dateStr) + " hui " + str(len(dateStr)) )
            else:
                commandStr = commandStr[0]
                dateStr = False

        if dateStr == False:
            dateToday = datetime.now()
            # print(dateToday)
            dayToday = dateToday.today().strftime('%d')  # '%d/%m/%Y'
            dateToday = dateToday.today().strftime('%d.%m.%Y')
            # dayToday = int(dayToday) * 1 #убираем 0
            dateTommorow = datetime.today() + timedelta(days=1)
            dateTommorow = dateTommorow.strftime('%d.%m.%Y')

            answerLineDateToday = "Сегодня " + dateToday + " на смене:"
            answerLineDateTommorow = "\n\nЗавтра " + dateTommorow + " на смене: "
        else:
            # if re.match(r’\d{1,2}\.\d{1,2}\.\d{2,4}$’, dateStr):
            dateToday = dateStr
            dateToday = datetime.strptime(dateToday, '%d.%m.%Y')
            dateTommorow = dateToday.date() + timedelta(days=1)
            dateTommorow = dateTommorow.strftime('%d.%m.%Y')
            dateToday = dateToday.strftime('%d.%m.%Y')

            answerLineDateToday = "*" + dateToday + "* на смене:"
            answerLineDateTommorow = "\n\n*" + dateTommorow + "* на смене: "

        # answerMess = "Техники РН, РЛ и связи:\n"
        # answerMess += "Сегодня " + dateToday + " на смене:"
        # wb = load_workbook('tech.xlsx')
        # sheet = wb.get_sheet_by_name('Апрель')
        
        ####
        #### >>>>>><<<<<<<<
        ####

        answerMess = "Техники РН, РЛ и связи:\n"
        # answerMess += "Сегодня " + dateToday + " на смене:" + dateStr
        answerMess += answerLineDateToday
        answerMess += "\nC ночи: " + getInfoSmena(dateToday, "tech", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateToday, "tech", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateToday, "tech", "3.5")

        # answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        answerMess += answerLineDateTommorow
        answerMess += "\nC ночи: " + getInfoSmena(dateTommorow, "tech", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateTommorow, "tech", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateTommorow, "tech", "3.5")

        answerMess += "\n\nСменные:\n"
        # answerMess += "Сегодня " + dateToday + " на смене:"
        answerMess += answerLineDateToday
        answerMess += "\nC ночи: " + getInfoSmena(dateToday, "ing", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateToday, "ing", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateToday, "ing", "3.5")

        # answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        answerMess += answerLineDateTommorow
        answerMess += "\nC ночи: " + getInfoSmena(dateTommorow, "ing", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateTommorow, "ing", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateTommorow, "ing", "3.5")

        # answerMess = markdown.markdown(answerMess)
        bot.send_message(answerUser, answerMess,  parse_mode="markdown") #message.chat.id
        # bot.send_message(message.from_user.id, answerMess)
    elif commandStr == "План на день" or commandStr == "/plan":
        dateToday = datetime.now()
        dayToday = dateToday.today().strftime('%Y-%m-%d')  # '%d/%m/%Y' 2022-04-01
        dayToday = dateToday.today().strftime('%Y-%m-%d')  # '%d/%m/%Y' 2022-04-01
        yearNow = dateToday.today().strftime('%Y')  # '%d/%m/%Y' 2022-04-01
        dayNumToday = dateToday.today().strftime('%w')
        try:
            connection = mysql.connector.connect(host='10.152.36.10',
                                                 database='arm_vdo',
                                                 user='root',
                                                 password='3333')

            answerMess = "План работ на " + dayToday + ": \n"
            ## TO2
            sql_select_Query = 'SELECT concat(ob.vidTO, " ", nameOb.name, " №", nameOb.numZav ), idSred FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and vidTO = "TO2" and dayTO = "' + dayNumToday + '" and nameOb.active = 1 '

            # sql_select_TO2 = 'SELECT concat(ob.vidTO, " ", nameOb.name) FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and ob.dayTO = "' + dayNumToday + '" and vidTO = "TO2" '
            cursor = connection.cursor()
            cursor.execute(sql_select_Query)
            records = cursor.fetchall()
            for row in records:
                idSred = str(row[1])
                sql_select_Query_Norma = 'SELECT norma, osnovanieTO FROM arm_vdo.grafTO_norma WHERE idSred = "'+ idSred +'" and yearTO = "' + yearNow + '" and vidTO = "TO2" '
                cursorNorma = connection.cursor()
                cursorNorma.execute(sql_select_Query_Norma)
                recordsNorma = cursorNorma.fetchone() 
                # print("{}".format(recordsNorma))
                
                
                cursorNorma.close()
                # if recordsOsnovanieTO 
                if recordsNorma is not None:
                    recordsOsnovanieTO = recordsNorma[1]
                    recordsNorma = recordsNorma[0]
                    
                    # print("{}".format(recordsOsnovanieTO))

                    if recordsNorma == '':
                        recordsNorma = 'нет информации'
                    else:
                        recordsNorma = math.ceil(float(recordsNorma) * 60)
                        recordsNorma = str(timedelta(seconds=recordsNorma * 60)) 
                        recordsNorma = str(recordsNorma)[:-3]
                        recordsNorma = recordsNorma[:1] + ' ч. ' + recordsNorma[2:] + ' мин.'
                    answerMess += row[0] + ' - ' + recordsNorma + '\n' + recordsOsnovanieTO + '\n\n'  # +  ' мин.\n' 
                    recordsNorma = 0
                else: 
                    answerMess += row[0] + ' - нет информации\n'
            # answerMess = answerMess.replace('&quot;','"')
            ## TO3-TO6
            sql_select_Query = 'SELECT concat(ob.vidTO, " ", nameOb.name, " №", nameOb.numZav), idSred, vidTO FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and datePlanTO = "' + dayToday + '" and active = 1'
            cursor = connection.cursor()
            cursor.execute(sql_select_Query)
            records = cursor.fetchall()
            for row in records:
                idSred = str(row[1])
                vidTO = str(row[2])
                # print  (idSred + ' vidTO ' + vidTO)
                sql_select_Query_Norma = 'SELECT norma, osnovanieTO FROM arm_vdo.grafTO_norma WHERE idSred = "'+ idSred +'" and vidTO = "'+ vidTO +'" and yearTO = "' + yearNow + '" '
                cursorNorma = connection.cursor()
                cursorNorma.execute(sql_select_Query_Norma)
                recordsNorma = cursorNorma.fetchone()
                
                # print("{}".format(recordsNorma))
                cursorNorma.close()
                if recordsNorma is not None:
                    recordsOsnovanieTO = recordsNorma[1]

                    print("{}".format(recordsOsnovanieTO))
                    recordsNorma = recordsNorma[0]
                    if recordsNorma == '':
                        recordsNorma = 'нет информации'
                    else:                            
                        recordsNorma = math.ceil(float(recordsNorma) * 60)
                        recordsNorma = str(timedelta(seconds=recordsNorma * 60)) 
                        recordsNorma = str(recordsNorma)[:-3]
                        recordsNorma = recordsNorma[:1] + ' ч. ' + recordsNorma[2:] + ' мин.'
                    answerMess += row[0] + ' - ' + recordsNorma + '\n' + recordsOsnovanieTO + '\n\n' #+  ' мин.\n'
                    recordsNorma = 0
                else: 
                    answerMess += row[0] + ' - нет информации\n'

                # print(row[0])
                # answerMess += row[0] + '\n'
            answerMess = answerMess.replace('&quot;', '"')
        except mysql.connector.Error as e:
            print("Error reading data from MySQL table", e)
        finally:
            if connection.is_connected():
                connection.close()
                cursor.close()
        # bot.send_message(message.chat.id, answerMess)
        bot.send_message(answerUser, answerMess)
    elif commandStr == "/zarplata":
        objects = ["Ццццц!!!", "Несссссссыыыыыы, она будет!!", "Опять ты",
                   "Может поработаем лучше, чем в телефоне залипать!", "Какая нахрен тебе зарплата! Работай давай!!!",
                   "Хули ты ноешь!! Иди трудись!!!", "Еще раз спросишь передам куда надо!!!", "Вот ты зануда",
                   "И чего с ней не так"]
        answerMess = objects[random.randint(0, len(objects)) - 1]
        bot.send_message(answerUser, answerMess) #message.chat.id
    elif commandStr == "/buhat":
        objects = ["Время? Место?", 'Сегодня, '+message.from_user.first_name+', на твои', "Только по пятницам", "Отличный вариант!!", "Пиво пиво водочка",
                   "Может в другой раз", "ЗОЖ наше все"]
        answerMess = objects[random.randint(0, len(objects)) - 1]
        bot.send_message(answerUser, answerMess) #message.chat.id
    elif commandStr == "Помощь" or commandStr == "/help":
        answerMess = "Информация по боту:"
        answerMess += "\n/smena - вывод информации по смене"
        answerMess += "\n/plan - план ТО на день"
        answerMess += "\n/pogoda - информация о погоде на текущее время по городу Салехард"
        bot.send_message(answerUser, answerMess)


# else:
#     bot.send_message(message.chat.id, message.text)
#     print(message.text)

# https://ru.stackoverflow.com/questions/711998/%D0%9D%D0%B5%D0%B4%D0%BE%D1%81%D1%82%D0%B0%D1%82%D0%BE%D0%BA-bot-polling
#1308468425
#-718406466 - Общий чат
# def send_messageToParty():
#     bot.send_photo(-718406466, photo=open('party1808.jpg', 'rb'))


def schedule_checker():
    while True:
        schedule.run_pending()
        sleep(1)

def function_to_run_1():
    print("This is a message to send.")
    return bot.send_message(1308468425, "Пора собираться на выход")

def function_to_run_2():
    print("This is a message to send.")
    return bot.send_message(1308468425, "Пора выходить на автобус")

def sendShedulerWeather():
    owm = pyowm.OWM('a58cd71e6d3057b4ce76c3da27076585')
    ob = owm.weather_at_place('Салехард, RU')
    w = ob.get_weather()
    temp = w.get_temperature('celsius')["temp"]
    wind = w.get_wind()["speed"]
    windGust = w.get_wind()["gust"]
    pressure = w.get_pressure()["press"] * 0.750063755419211  # / 1.333
    answerMess = "В Салехарде сейчас:\n\nтемпература: " + str(temp) + "°С\nветер: " + str(
        wind) + " м/с\nпорывы до: " + str(windGust) + " м/с\nдавление: " + str(round(pressure, 1)) + " mmHg"
    # return bot.send_message(1308468425, answerMess)
    bot.send_message(1308468425, answerMess)

def sendShedulerLianaDiscSpace():
    percent = os.popen("sshpass -p otstoj111 ssh -p 10018 user@10.3.103.100 df / | grep / | awk '{ print $5}' | sed 's/%//g'").read()
    if int(percent) >= 90:
        answerMess = "!!! ВНИМАНИЕ !!!\n\nВ корневом разделе сервера ЛИАНА, осталось слишком мало дискового пространства.\nИспользуется: " + str(percent)
        bot.send_message(1308468425, answerMess)
        bot.send_message(609169556, answerMess) #Виктор Борисович
    # print('result: ', percent.read())


def remindJob():
    ## функиция напоминания о выходе на работу
    ## напоминаем о смене на текущий вечер и день завтра утро
    dateToday = datetime.now()  
    dateToday = dateToday.today().strftime('%d.%m.%Y')
    dateTommorow = datetime.today() + timedelta(days=1)
    dateTommorow = dateTommorow.strftime('%d.%m.%Y')
    answerMess = "\n!!! Напоминание !!!"
    answerMess += "\n\nВ ночь:"
    answerMess += "\nтехник - " + getInfoSmena(dateToday, "tech", "3.5")
    answerMess += "\nсм. инж - " + getInfoSmena(dateToday, "ing", "3.5")
    answerMess += "\n\nЗавтра день:"
    answerMess += "\nтехник - " + getInfoSmena(dateTommorow, "tech", "12")
    answerMess += "\nсм. инж - " + getInfoSmena(dateTommorow, "ing", "12")
    answerMess += "\n\nИнформация справочная, лучше уточнять"
    bot.send_message(-1001651843114, answerMess)
    # bot.send_message(1308468425, answerMess)

if __name__ == "__main__": 
    schedule.every().day.at("12:12:12").do(remindJob)
    # schedule.every().day.at("07:45").do(function_to_run_1)
    # schedule.every().day.at("07:50").do(function_to_run_2)
    schedule.every().day.at("06:30:33").do(sendShedulerWeather)
    schedule.every().day.at("09:09:09").do(sendShedulerLianaDiscSpace)
    Thread(target=schedule_checker).start()  

    while True:
        # ScheduleMessage.start_process()
        try:
            bot.polling(none_stop=True) 
        except Exception as e:
            # logger.error(e)  # или просто print(e) если у вас логгера нет,
            print(e)
            # или import traceback; traceback.print_exc() для печати полной инфы
            time.sleep(15)



# {
#     "coord":{
#         # "lon":-0.13,
#         "lat":51.51
#         },
#     "weather":[
#         {"id":300,
#          "main":"Drizzle",
#          "description":"light intensity drizzle",
#          "icon":"09d"
#          }
#         ],
#     "base":"stations",
#     "main":{
#         "temp":280.32,
#         "pressure":1012,
#         "humidity":81,
#         "temp_min":279.15,
#         "temp_max":281.15
#         },
#     "visibility":10000,
#     "wind":{
#         "speed":4.1,
#         "deg":80
#         },
#     "clouds":{
#         "all":90
#         },
#     "dt":1485789600,
#     "sys":{
#         "type":1,
#         "id":5091,
#         "message":0.0103,
#         "country":"GB",
#         "sunrise":1485762037,
#         "sunset":1485794875
#         },
#     "id":2643743,
#     "name":"London",
#     "cod":200
# }

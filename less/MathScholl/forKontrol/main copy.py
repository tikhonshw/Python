from reportlab.lib.units import inch, cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import matplotlib.pyplot as plt 
import textwrap
 

import openpyxl
import random
from os import path as dirPath
 
path = dirPath.dirname('/home/user/MySoft/Python/less/MathScholl/forKontrol/main.py')
print(path)

def createPDF(path,nameFile):
    canvas = Canvas(path + '/' + nameFile + '.pdf')
    return canvas

def genPDF(path,canvas,x,y,primer):
    # canvas = Canvas(path + '/printLabel.pdf', pagesize=(5.8 * cm, 4 * cm))
    pdfmetrics.registerFont(TTFont('Arial', path + '/ArialCyr.ttf')) 
    canvas.setFont("Arial", 14)
    canvas.drawString(y * cm, x * cm, "{}".format(primer)) 

def savePDF(canvas):
    canvas.save()

wb = openpyxl.Workbook()

# Удаление листа, создаваемого по умолчанию, при создании документа
for sheet_name in wb.sheetnames:
    sheet = wb.get_sheet_by_name(sheet_name)
    wb.remove_sheet(sheet)
# sheet = wb.active  

listDown = 1000
maxGenPrimer = 200 * listDown


sheet = wb.create_sheet("Деление без остатка х х")
canvas = createPDF(path,'Деление без остатка х х')
rCount = 1
cCount = 1
countDel = 0
xStart = 1.5
yStart = 2.5
x = xStart + 0.7
y = yStart
for i in range(0, maxGenPrimer):
    per1 = random.randint(2,9)
    per2 = random.randint(2,9)
    if per1 % per2 == 0: 
        print(i, per1, ':', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}:{}=".format(per1,per2) 
        primer = "{}:{}=".format(per1,per2)
        genPDF(path,canvas,x,y,primer)
    else:
        while per1 % per2 != 0:
            if per1 % per2 == 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
                primer = "{}:{}=".format(per1,per2)
                genPDF(path,canvas,x,y,primer)
            else:
                per1 = random.randint(2,9)
                per2 = random.randint(2,9)
                print(i, '---===---', per1, ':', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
                if per1 % per2 == 0: 
                    primer = "{}:{}=".format(per1,per2)
                    genPDF(path,canvas,x,y,primer)
    if (i + 1) % 10 == 0:
        canvas.line(yStart * cm, (x + 0.525) * cm, (yStart + 17.5) *cm, (x + 0.525) * cm)

    if x >= 27:
        y = y + 3.5
        x = xStart
        if y > 17:
            canvas.showPage()
            y = yStart
            x = xStart
    x = x + 0.65

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1
savePDF(canvas)
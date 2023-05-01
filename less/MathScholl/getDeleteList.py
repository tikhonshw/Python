from reportlab.lib.units import inch, cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import matplotlib.pyplot as plt 
import textwrap
 


import openpyxl
import random
from os import path as dirPath
 
path = dirPath.dirname('/home/user/MySoft/Python/less/MathScholl/myList.xlsx')
print(path)

def createPDF(path,nameFile):
    canvas = Canvas(path + '/' + nameFile + '.pdf')
    return canvas

def genPDF(path,canvas,x,y,primer):
    # canvas = Canvas(path + '/printLabel.pdf', pagesize=(5.8 * cm, 4 * cm))
    pdfmetrics.registerFont(TTFont('Arial', path + '/ArialCyr.ttf')) 
    canvas.setFont("Arial", 14)
    canvas.drawString(y * cm, x * cm, "{}".format(primer)) 

def savePDF(canvas):
    canvas.save()

wb = openpyxl.Workbook()

# Удаление листа, создаваемого по умолчанию, при создании документа
for sheet_name in wb.sheetnames:
    sheet = wb.get_sheet_by_name(sheet_name)
    wb.remove_sheet(sheet)
# sheet = wb.active 

# c1 = sheet.cell(row = 1, column = 1) 
# c1.value = "ANKIT"
# c2 = sheet.cell(row= 1 , column = 2)
# c2.value = "RAI" 
# c3 = sheet['A2']
# c3.value = "RAHUL" 
# c4 = sheet['B2']
# c4.value = "RAI" 

maxGenPrimer = 200 * 3
listDown = 1

sheet = wb.create_sheet("Деление без остатка х х")
canvas = createPDF(path,'DelNotOstX_X')

rCount = 1
cCount = 1
countDel = 0
xStart = 1.5
yStart = 2.5
x = xStart + 0.7
y = yStart
for i in range(0, maxGenPrimer):
    per1 = random.randint(2,9)
    per2 = random.randint(2,9)
    if per1 % per2 == 0: 
        print(i, per1, ':', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}:{}=".format(per1,per2) 
        primer = "{}:{}=".format(per1,per2)
        genPDF(path,canvas,x,y,primer)
    else:
        while per1 % per2 != 0:
            if per1 % per2 == 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
                primer = "{}:{}=".format(per1,per2)
                genPDF(path,canvas,x,y,primer)
            else:
                per1 = random.randint(2,9)
                per2 = random.randint(2,9)
                print(i, '---===---', per1, ':', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
                if per1 % per2 == 0: 
                    primer = "{}:{}=".format(per1,per2)
                    genPDF(path,canvas,x,y,str(i) + ' ' + primer)
    if (i + 1) % 10 == 0:
        canvas.line(yStart * cm, (x + 0.525) * cm, (yStart + 17.5) *cm, (x + 0.525) * cm)

    if x >= 27:
        y = y + 3.5
        x = xStart
        if y > 17:
            canvas.showPage()
            y = yStart
            x = xStart
    x = x + 0.65

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1
savePDF(canvas)

sheet = wb.create_sheet('Деление без остатка хх х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(2,9)
    if per1 % per2 == 0: 
        print(i, per1, ':', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}:{}=".format(per1,per2) 
    else:
        while per1 % per2 != 0:
            if per1 % per2 == 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(2,9)
                print(i, '---===---', per1, ':', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}:{}=".format(per1,per2)
                 
    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Умножение x х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(1,9) 

    print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}*{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1


sheet = wb.create_sheet('Сложение х+х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(2,9)
    per2 = random.randint(2,9)
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сложение хх+х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(2,9)
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

# sheet = wb.create_sheet('Сложение хх+хх > 100')
# rCount = 1
# cCount = 1
# countDel = 0
# for i in range(0, 10000):
#     per1 = random.randint(11,99)
#     per2 = random.randint(11,99)
#     c = sheet.cell(row = rCount, column = cCount) 
#     c.value = "{}+{}=".format(per1,per2)

#     if rCount == 46 * 100:
#         cCount = cCount + 2
#         rCount = 0
#     rCount = rCount + 1

sheet = wb.create_sheet('Сложение хх+х < 100')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(2,9) 
    if per1 + per2 <= 100: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}+{}=".format(per1,per2)
    else:
        while per1 + per2 > 100:
            if per1 + per2 < 100: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(2,9)
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сложение хх+хx < 100')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(11,99) 
    if per1 + per2 <= 100: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}+{}=".format(per1,per2)
    else:
        while per1 + per2 > 100:
            if per1 + per2 < 100: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(11,99) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сложение хх+х > 100')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(2,9) 
    if per1 + per2 >= 100: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}+{}=".format(per1,per2)
    else:
        while per1 + per2 < 100:
            if per1 + per2 > 100: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(2,9) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сложение хх+хx > 100')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(11,99) 
    if per1 + per2 >= 100: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}+{}=".format(per1,per2)
    else:
        while per1 + per2 < 100:
            if per1 + per2 > 100: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(11,99) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 / per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}+{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Вычетание х-х < 10')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(1,9) 
    if per1 - per2 >= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 < 0:
            if per1 - per2 > 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(1,9)
                per2 = random.randint(1,9) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Вычетание хx-х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(10,99)
    per2 = random.randint(1,9) 
    if per1 - per2 >= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 < 0:
            if per1 - per2 > 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(10,99)
                per2 = random.randint(1,9) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Вычетание хx-хx > 0')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(10,99)
    per2 = random.randint(10,99) 
    if per1 - per2 >= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 < 0:
            if per1 - per2 > 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(10,99)
                per2 = random.randint(10,99) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1


sheet = wb.create_sheet('Вычетание x-х < 0')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(1,9) 
    if per1 - per2 <= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 > 0:
            if per1 - per2 < 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(1,9)
                per2 = random.randint(1,9) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Вычетание x-хx < 0')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(11,99) 
    if per1 - per2 <= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 > 0:
            if per1 - per2 < 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(1,9)
                per2 = random.randint(11,99) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Вычетание xx-хx < 0')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(11,99) 
    if per1 - per2 <= 0: 
        print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
        c = sheet.cell(row = rCount, column = cCount) 
        c.value = "{}-{}=".format(per1,per2)
    else:
        while per1 - per2 > 0:
            if per1 - per2 < 0: 
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)
            else:
                per1 = random.randint(11,99)
                per2 = random.randint(11,99) 
                print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
                c = sheet.cell(row = rCount, column = cCount) 
                c.value = "{}-{}=".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сравнение x х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(1,9) 

    print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}___{}".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сравнение xх хх')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(11,99) 

    print(i, '---===---', per1, '+', per2 , '=', str(int(per1 - per2)))
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}___{}".format(per1,per2)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сравнение x_х  x_х')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(1,9)
    per2 = random.randint(1,9) 
    per3 = random.randint(1,9)
    per4 = random.randint(1,9)  
    arrTmp = ['+','-']
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}{}{}___{}{}{}".format(per1,arrTmp[random.randint(0,1)],per2,per3,arrTmp[random.randint(0,1)],per4)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1

sheet = wb.create_sheet('Сравнение xx_хx  xx_хx')
rCount = 1
cCount = 1
countDel = 0
for i in range(0, maxGenPrimer):
    per1 = random.randint(11,99)
    per2 = random.randint(11,99) 
    per3 = random.randint(11,99)
    per4 = random.randint(11,99)  
    arrTmp = ['+','-']
    c = sheet.cell(row = rCount, column = cCount) 
    c.value = "{}{}{}___{}{}{}".format(per1,arrTmp[random.randint(0,1)],per2,per3,arrTmp[random.randint(0,1)],per4)

    if rCount == 46 * listDown:
        cCount = cCount + 2
        rCount = 0
    rCount = rCount + 1



wb.save(path + '/myList.xlsx')

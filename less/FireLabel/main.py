from reportlab.lib.units import inch, cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import matplotlib.pyplot as plt
from datetime import datetime
import textwrap
import os
from os import path as dirPath

path = dirPath.dirname('/home/user/MySoft/Python/less/FireLabel/main.py')
print(path)

canvas = Canvas(path + '/printLabel.pdf', pagesize=(5.8 * cm, 4 * cm))

pdfmetrics.registerFont(TTFont('Arial', path + '/ArialCyr.ttf'))
pdfmetrics.registerFont(TTFont('ArialBold', path + '/ArialBold.ttf'))

# для обработки xlsx https://pythonist.ru/kak-chitat-excel-fajl-xlsx-v-python/
from openpyxl import load_workbook
wb = load_workbook(path + '/base.xlsx')
sheet = wb.get_sheet_by_name('учет')
fio = sheet.cell(row=1, column=5).value
dolj = sheet.cell(row=3, column=5).value

strAll = []
str = []
for i in range(4, sheet.max_row):
    str = []
    for col in sheet.iter_cols(1, sheet.max_column):
        str.append(col[i].value)
    strAll.append(str)

circle1 = plt.Circle((0, 0), 0.2, color='r')

dateToday = datetime.now()
dateToday = dateToday.today().strftime('%d.%m.%Y')

for num, model, ves, dav, mesto in strAll:
    canvas.circle(5.3 * cm, 3.5 * cm, 10, stroke=1, fill=0)
    if num < 10: 
        canvas.drawString(5.18 * cm, 3.35 * cm, "{}".format(num)) 
    else:
        canvas.drawString(5.06 * cm, 3.35 * cm, "{}".format(num))
    
    canvas.setFont("ArialBold", 14)
    canvas.drawString(1.35 * cm, 3.35 * cm, "{}".format(dateToday))

    canvas.setFont("ArialBold", 10)
    canvas.drawString(0.3 * cm, 2.9 * cm, "Модель:") 
    canvas.setFont("Arial", 10)
    canvas.drawString(2 * cm, 2.9 * cm, "{}".format(model)) 

    canvas.setFont("ArialBold", 10)
    canvas.drawString(0.3 * cm, 2.5 * cm, "Расположение:") 
    canvas.setFont("Arial", 10)
    x = 15
    y = 60
    if len(mesto) > 25:
        wrap_text = textwrap.wrap(mesto, width=25)
        for i in range(0, len(wrap_text)):
            canvas.drawString(x, y, "{}".format(wrap_text[i] ) )
            y = y - 8
    else:
        canvas.drawString(x, y, mesto)
        y = y - 8 
    canvas.setFont("ArialBold", 10)
    canvas.drawString(0.3 * cm, y, "Вес:")
    canvas.setFont("Arial", 10)
    canvas.drawString(1.2 * cm, y, "{}".format(ves ) )
    canvas.setFont("ArialBold", 10)
    if dav is not None:
        canvas.drawString(2.9 * cm, y, "Давление:")
        canvas.setFont("Arial", 10)
        canvas.drawString(4.9 * cm, y, "{}".format(dav ) )
    y = y - 10 
    canvas.setFont("ArialBold", 10)
    canvas.drawString(0.3 * cm, y, "Ответственный: {}".format(fio))
    y = y - 8 
    canvas.setFont("Arial", 8)
    canvas.drawString(0.3 * cm, y, "{}".format(dolj))

    canvas.line(0.3*cm,0.3*cm,5.5*cm,0.3*cm)
    canvas.showPage()
canvas.save()

os.system('echo "Наклейки для огнетушителей\n' + '" | mutt -s "Наклейки для огнетушителей" -e "my_hdr From: GoodJob <tixon@yamal.ans.aero>" -a ' + path + '/printLabel.pdf  -- kdp@yamal.ans.aero, tixon@yamal.ans.aero') #  energetik@yamal.ans.aero,    kdp@yamal.ans.aero, laz_kdp@yamal.ans.aero, 
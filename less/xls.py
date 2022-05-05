# //https://codecamp.ru/blog/python-excel-tutorial/
# import xlwings as xw
# import pandas as pd
# from forex_python.converter import CurrencyRates
# from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
from datetime import datetime #https://pythonworld.ru/moduli/modul-datetime.html

dateToday = datetime.now()
dayToday = dateToday.today().strftime('%d') #'%d/%m/%Y'
dayToday = int(dayToday) * 1 #убираем 0
# print( str(dayToday) )

# Import `load_workbook` module from `openpyxl`


# Load in the workbook
wb = load_workbook('01.xlsx')

sheet = wb.get_sheet_by_name('Апрель')
# print(sheet['A15'].value)

for i in range(15, 30):
    if sheet.cell(row=i, column=dayToday + 1).value is not None:
        # print(i, sheet.cell(row=i, column=1).value)
        workDay = sheet.cell(row=i, column=2).value
        if workDay == int(12):
            print(i, sheet.cell(row = i - 1, column=1).value + "День 12")
        if workDay == float(3.5):
            print(i, sheet.cell(row = i - 1, column=1).value + "Ночь 3.5")
        if workDay == float(8.5):
            print(i, sheet.cell(row = i - 1, column=1).value + "С ночной 8.5")
    if sheet.cell(row = i, column = dayToday + 2).value is not None:
        # print(i, sheet.cell(row=i, column=1).value)
        workDay = sheet.cell(row=i, column= dayToday + 2).value
        if workDay == int(12):
            print(i, sheet.cell(row = i - 1, column=1).value + "завтра День 12")
        if workDay == float(3.5):
            print(i, sheet.cell(row = i - 1, column=1).value + "завтра Ночь 3.5")
        if workDay == float(8.5):
            print(i, sheet.cell(row = i - 1, column=1).value + "завтра С ночной 8.5")

# Get sheet names
# print(wb.get_sheet_names())

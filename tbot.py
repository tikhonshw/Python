import telebot
import pyowm
import random
import time
# from datetime import datetime
from datetime import datetime, timedelta
# import timedelta
import mysql.connector

# для обработки xlsx
from openpyxl import load_workbook

bot = telebot.TeleBot("5125341770:AAF11nLzMCoeFV-gf96iL19hDyhOfidqo7g")


#
# from pyowm import OWM
# owm = OWM('a58cd71e6d3057b4ce76c3da27076585')
# mgr = owm.weather_manager()
# observation = mgr.weather_at_place('Москва')
# w = observation.weather
# print(w, w.wind())


@bot.message_handler(content_types=['text'])
def send_echo(message):
    #    if message.text == "/pogoda":
    #        print('/pogoda Start')
    # owm = OWM('a58cd71e6d3057b4ce76c3da27076585')
    # mgr = owm.weather_manager()
    if message.text == "/pogoda":
        owm = pyowm.OWM('a58cd71e6d3057b4ce76c3da27076585')
        ob = owm.weather_at_place('Салехард, RU')
        w = ob.get_weather()

        temp = w.get_temperature('celsius')["temp"]
        wind = w.get_wind()["speed"]
        windGust = w.get_wind()["gust"]
        pressure = w.get_pressure()["press"] * 0.750063755419211  # / 1.333
        answerMess = "В Салехарде сейчас:\n\nтемепература: " + str(temp) + "°С\nветер: " + str(
            wind) + " м/с\nпорывы до: " + str(windGust) + " м/с\nдавление: " + str(round(pressure, 1)) + " mmHg"
        bot.send_message(message.chat.id, answerMess)

    elif message.text == "/smena":
        dateToday = datetime.now()
        dayToday = dateToday.today().strftime('%d')  # '%d/%m/%Y'
        dateToday = dateToday.today().strftime('%d.%m.%Y')
        # dayToday = int(dayToday) * 1 #убираем 0
        dateTommorow = datetime.today() + timedelta(days=1)
        dateTommorow = dateTommorow.strftime('%d.%m.%Y')

        # answerMess = "Техники РН, РЛ и связи:\n"
        # answerMess += "Сегодня " + dateToday + " на смене:"
        # wb = load_workbook('tech.xlsx')
        # sheet = wb.get_sheet_by_name('Апрель')
        def getInfoSmena(dateInfo, dolgnost, smena):
            ddmmYYYY = dateInfo.split(".")
            dd = int(ddmmYYYY[0]) * 1
            mm = int(ddmmYYYY[1]) * 1
            massMonth = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь",
                         "Октябрь", "Ноябрь", "Декабрь"]
            if dolgnost == 'tech':
                wb = load_workbook('tech.xlsx')  # подключаем файл график техников
            if dolgnost == 'ing':
                wb = load_workbook('ing.xlsx')  # подключаем файл график техников
            sheet = wb.get_sheet_by_name(massMonth[mm])
            for i in range(15, 32):
                if sheet.cell(row=i, column=dd + 1).value is not None:
                    workDay = sheet.cell(row=i, column=dd + 1).value
                    if str(smena) == str(workDay):
                        return sheet.cell(row=i - 1, column=1).value

        answerMess = "Техники РН, РЛ и связи:\n"
        answerMess += "Сегодня " + dateToday + " на смене:"
        answerMess += "\nC ночи: " + getInfoSmena(dateToday, "tech", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateToday, "tech", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateToday, "tech", "3.5")

        answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        answerMess += "\nC ночи: " + getInfoSmena(dateTommorow, "tech", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateTommorow, "tech", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateTommorow, "tech", "3.5")

        answerMess += "\n\nСменные:\n"
        answerMess += "Сегодня " + dateToday + " на смене:"
        answerMess += "\nC ночи: " + getInfoSmena(dateToday, "ing", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateToday, "ing", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateToday, "ing", "3.5")

        answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        answerMess += "\nC ночи: " + getInfoSmena(dateTommorow, "ing", "8.5")
        answerMess += "\nВ день: " + getInfoSmena(dateTommorow, "ing", "12")
        answerMess += "\nВ ночь: " + getInfoSmena(dateTommorow, "ing", "3.5")

        # for i in range(15, 32):
        #     if sheet.cell(row=i, column = dayToday + 1).value is not None:
        #         # print(i, sheet.cell(row=i, column=1).value)
        #         workDay = sheet.cell(row=i, column = dayToday + 1).value
        #         if workDay == int(12):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "День 12")
        #             answerMess += '\nВ день: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(3.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "Ночь 3.5")
        #             answerMess += '\nВ ночь: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(8.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "С ночной 8.5")
        #             answerMeiss += '\nС ночи: ' + sheet.cell(row = i - 1, column=1).value
        # answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        # for i in range(15, 32):
        #     if sheet.cell(row = i, column = dayToday + 2).value is not None:
        #         # print(i, sheet.cell(row=i, column=1).value)
        #         workDay = sheet.cell(row=i, column = dayToday + 2).value
        #         if workDay == int(12):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "День 12")
        #             answerMess += '\nВ день: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(3.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "Ночь 3.5")
        #             answerMess += '\nВ ночь: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(8.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "С ночной 8.5")
        #             answerMess += '\nС ночи: ' + sheet.cell(row = i - 1, column=1).value
        # answerMess += "\n\nСменные:\n"
        # answerMess += "Сегодня " + dateToday + " на смене:"
        # wb = load_workbook('ing.xlsx')
        # sheet = wb.get_sheet_by_name('Апрель')
        # for i in range(15, 32):
        #     if sheet.cell(row=i, column = dayToday + 1).value is not None:
        #         # print(i, sheet.cell(row=i, column=1).value)
        #         workDay = sheet.cell(row=i, column = dayToday + 1).value
        #         if workDay == int(12):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "День 12")
        #             answerMess += '\nВ день: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(3.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "Ночь 3.5")
        #             answerMess += '\nВ ночь: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(8.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "С ночной 8.5")
        #             answerMess += '\nС ночи: ' + sheet.cell(row = i - 1, column=1).value
        # answerMess += "\n\nЗавтра " + dateTommorow + " на смене: "
        # for i in range(15, 32):
        #     if sheet.cell(row = i, column = dayToday + 2).value is not None:
        #         # print(i, sheet.cell(row=i, column=1).value)
        #         workDay = sheet.cell(row=i, column = dayToday + 2).value
        #         if workDay == int(12):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "День 12")
        #             answerMess += '\nВ день: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(3.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "Ночь 3.5")
        #             answerMess += '\nВ ночь: ' + sheet.cell(row = i - 1, column=1).value
        #         if workDay == float(8.5):
        #             # print(i, sheet.cell(row = i - 1, column=1).value + "С ночной 8.5")
        #             answerMess += '\nС ночи: ' + sheet.cell(row = i - 1, column=1).value

        bot.send_message(message.chat.id, answerMess)
    elif message.text == "/plan":
        dateToday = datetime.now()
        dayToday = dateToday.today().strftime('%Y-%m-%d')  # '%d/%m/%Y' 2022-04-01
        dayToday = dateToday.today().strftime('%Y-%m-%d')  # '%d/%m/%Y' 2022-04-01
        yearNow = dateToday.today().strftime('%Y')  # '%d/%m/%Y' 2022-04-01
        dayNumToday = dateToday.today().strftime('%w')
        try:
            connection = mysql.connector.connect(host='10.152.36.10',
                                                 database='arm_vdo',
                                                 user='root',
                                                 password='3333')

            answerMess = "План работ на " + dayToday + ": \n"
            ## TO2
            sql_select_Query = 'SELECT concat(ob.vidTO, " ", nameOb.name) FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and vidTO = "TO2" and dayTO = "' + dayNumToday + '" and nameOb.active = 1 '

            # sql_select_TO2 = 'SELECT concat(ob.vidTO, " ", nameOb.name) FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and ob.dayTO = "' + dayNumToday + '" and vidTO = "TO2" '
            cursor = connection.cursor()
            cursor.execute(sql_select_Query)
            records = cursor.fetchall()
            for row in records:
                # print(row[0])
                answerMess += row[0] + '\n'
            # answerMess = answerMess.replace('&quot;','"')
            ## TO3-TO6
            sql_select_Query = 'SELECT concat(ob.vidTO, " ", nameOb.name) FROM arm_vdo.grafTO_planTO as ob Left JOIN arm_vdo.grafTO_oborudovanie as nameOb ON ob.idSred = nameOb.id where yearTO = "' + yearNow + '" and datePlanTO = "' + dayToday + '"'
            cursor = connection.cursor()
            cursor.execute(sql_select_Query)
            records = cursor.fetchall()
            for row in records:
                # print(row[0])
                answerMess += row[0] + '\n'
            answerMess = answerMess.replace('&quot;', '"')
        except mysql.connector.Error as e:
            print("Error reading data from MySQL table", e)
        finally:
            if connection.is_connected():
                connection.close()
                cursor.close()
        # bot.send_message(message.chat.id, answerMess)
        bot.send_message(message.chat.id, answerMess)
    elif message.text == "/zarplata":
        objects = ["Ццццц!!!", "Несссссссыыыыыы, она будет!!", "Опять ты",
                   "Может поработаем лучше, чем в телефоне залипать!", "Какая нахрен тебе зарплата! Работай давай!!!",
                   "Хули ты ноешь!! Иди трудись!!!", "Еще раз спросишь передам куда надо!!!", "Вот ты зануда",
                   "И чего с ней не так"]
        answerMess = objects[random.randint(0, len(objects)) - 1]
        bot.send_message(message.chat.id, answerMess)
    elif message.text == "/buhat":
        objects = ["Время? Место?", "Сегодня на твои", "Только по пятницам", "Отличный вариант!!", "Пиво пиво водочка",
                   "Может в другой раз", "ЗОЖ наше все"]
        answerMess = objects[random.randint(0, len(objects)) - 1]
        bot.send_message(message.chat.id, answerMess)
    elif message.text == "/help":
        answerMess = "Информация по боту:"
        answerMess += "\n/smena - вывод информации по смене"
        answerMess += "\n/plan - план ТО на день"
        answerMess += "\n/pogoda - информация о погоде на текущее время по городу Салехард"
        bot.send_message(message.chat.id, answerMess)


# else:
#     bot.send_message(message.chat.id, message.text)
#     print(message.text)


# bot.polling( none_stop = True )
# https://ru.stackoverflow.com/questions/711998/%D0%9D%D0%B5%D0%B4%D0%BE%D1%81%D1%82%D0%B0%D1%82%D0%BE%D0%BA-bot-polling
while True:
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        # logger.error(e)  # или просто print(e) если у вас логгера нет,
        print(e)
        # или import traceback; traceback.print_exc() для печати полной инфы
        time.sleep(15)
